import time
import typing as t

from openpyxl import load_workbook
from datetime import datetime

from app.core.config import settings
from app.db.session import SessionLocal

from app import crud, models, schemas

db = SessionLocal()


def get_file(file="reklaliste.xlsx") -> load_workbook:
    wb_file = load_workbook(str(settings.SERVER_BASE_DIR) + '/' + file)
    i = 0
    for sheet in wb_file.sheetnames:
        i += 1
        if sheet.lower() == "liste wertreklas":
            break
        wb_file.active = i
    return wb_file.active


def fetch_claim_data(row: t.List) -> t.List:
    data = []
    for cell in row:
        data.append(cell)
    return data


def create_claim(*, data: t.List) -> t.Union[schemas.claim.ClaimBase, None]:
    claim_in = schemas.claim.ClaimBase()

    if isinstance(data[1], float):
        contract_nr = str(data[1]).split(".0")[0]
        claim_in.contract_nr = int(contract_nr)
    if isinstance(data[2], float) or isinstance(data[2], int):
        claim_in.discharge = float(data[2])
    if isinstance(data[3], float) or isinstance(data[3], int):
        claim_in.bill = float(data[3])
    if isinstance(data[7], datetime):
        claim_in.created_at = datetime.date(data[7])
    if isinstance(data[11], str):
        claim_in.kind = data[11].upper()
        try:
            claim_in.kind = models.ClaimKind(claim_in.kind).name
        except ValueError:
            claim_in.kind = None

    if claim_in.contract_nr is not None and claim_in.bill is not None and claim_in.created_at is not None:
        return claim_in
    return None


def iter_claim_rows(file: load_workbook) -> t.List[schemas.Claim]:
    i = 0
    for row in file.iter_rows(values_only=True, max_col=15):
        i += 1
        if i % 100 == 1:
            time.sleep(5)
        data = fetch_claim_data(row)
        echo = create_claim(data=data)
        if echo is not None:
            yield echo


def progress():
    rows = iter_claim_rows(file=get_file())
    for row in rows:
        ticket_exist = crud.ticket.get_by_kwargs(db, contract_nr=row.contract_nr,
                                                 kind=models.TicketKind.COMMISSION).first()
        if ticket_exist:
            # UPDATE CLAIM DATA
            row.store_internal_id = ticket_exist.store_internal_id
            row.owner_id = ticket_exist.owner_id
            row.ticket_id = ticket_exist.ticket_id

            print("...ticket ->", ticket_exist.ticket_id)
            claim_exist = crud.claim.get_by_kwargs(db, contract_nr=row.contract_nr, created_at=row.created_at, bill=row.bill).first()
            if not claim_exist:
                print("...claim not exist!")
                claim = crud.claim.create(db, obj_in=row)
                if claim:
                    print(f"...created new claim {claim.id}")
            else:
                print("...claim exist!")
                claim = crud.claim.update(db, db_obj=claim_exist, obj_in=row)
                if claim:
                    print(f"...updated claim {claim.id}")




